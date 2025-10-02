"""
Generador de archivos Excel con análisis de hipotecas.
"""

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .calculator import MortgageCalculator
from .models import MortgageData, MortgageResults


class ExcelGenerator:
    """Generador de reportes Excel para análisis de hipotecas."""

    def __init__(self, mortgage_data: MortgageData):
        self.data = mortgage_data
        self.calculator = MortgageCalculator(mortgage_data)
        self.results: Optional[MortgageResults] = None

    def generate_report(self, output_path: str = "analisis_hipoteca.xlsx") -> str:
        """
        Genera el reporte completo en Excel.

        Args:
            output_path: Ruta donde guardar el archivo Excel

        Returns:
            Ruta del archivo generado
        """
        # Realizar cálculos
        self.results = self.calculator.calculate()

        # Crear el archivo Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._create_input_sheet(writer)
            self._create_summary_sheet(writer)
            self._create_comparison_sheet(writer)
            self._create_amortization_sheet(writer, with_bonus=False)
            self._create_amortization_sheet(writer, with_bonus=True)
            self._create_bonus_analysis_sheet(writer)
            self._create_insurance_individual_analysis_sheet(writer)

        # Aplicar formato y fórmulas
        self._apply_formatting(output_path)
        self._add_formulas_to_sheets(output_path)

        return str(Path(output_path).absolute())

    def _create_input_sheet(self, writer: pd.ExcelWriter):
        """Crea la hoja con los datos de entrada."""
        data = {
            "Concepto": [
                "▼ DATOS DE LA HIPOTECA",
                "Capital prestado (€)",
                "Tasa de interés anual (%)",
                "Plazo (años)",
                "",
                "▼ BONIFICACIONES",
                "Bonificación por nómina (%)",
                "Bonificación por seguro de vida (%)",
                "Bonificación por seguro de hogar (%)",
                "Bonificación por tarjeta (%)",
                "Otras bonificaciones (%)",
                "",
                "▼ COSTES DE BONIFICACIONES",
                "Coste mensual seguro de vida (€)",
                "Coste mensual seguro de hogar (€)",
                "Cuota anual de la tarjeta (€)",
                "Otros costes mensuales (€)",
            ],
            "Valor": [
                "",
                self.data.capital,
                self.data.interest_rate,
                self.data.years,
                "",
                "",
                self.data.payroll_bonus,
                self.data.life_insurance_bonus,
                self.data.home_insurance_bonus,
                self.data.card_bonus,
                self.data.other_bonus,
                "",
                "",
                self.data.life_insurance_cost_monthly,
                self.data.home_insurance_cost_monthly,
                self.data.card_annual_fee,
                self.data.other_costs_monthly,
            ],
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Datos de Entrada", index=False)

    def _create_summary_sheet(self, writer: pd.ExcelWriter):
        """Crea la hoja resumen con fórmulas dinámicas."""
        # Esta hoja se llenará completamente con fórmulas en _add_formulas_to_sheets
        data = {
            "Concepto": [
                "▼ CÁLCULOS AUTOMÁTICOS",
                "Meses totales",
                "Bonificación total (%)",
                "Tipo efectivo sin bonif. (%)",
                "Tipo efectivo con bonif. (%)",
                "Cuota mensual SIN bonificaciones (€)",
                "Cuota mensual CON bonificaciones (€)",
                "Total a pagar SIN bonificaciones (€)",
                "Total a pagar CON bonificaciones (€)",
                "Intereses SIN bonificaciones (€)",
                "Intereses CON bonificaciones (€)",
                "Costes bonificaciones (€)",
                "Ahorro en intereses (€)",
                "Ahorro real (€)",
                "¿Vale la pena?",
                "Porcentaje de ahorro (%)",
            ],
            "Fórmula/Valor": [""] * 16,  # Se llenarán con fórmulas
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Resumen", index=False)

    def _create_comparison_sheet(self, writer: pd.ExcelWriter):
        """Crea una hoja de comparación detallada."""
        if not self.results:
            return

        data = {
            "Concepto": [
                "Capital prestado",
                "Tasa de interés",
                "Bonificaciones aplicadas",
                "Tasa con bonificaciones",
                "Plazo (meses)",
                "",
                "Cuota mensual",
                "Total intereses",
                "Total a pagar",
                "Costes bonificaciones",
                "Coste real total",
            ],
            "Sin Bonificaciones": [
                f"{self.data.capital:,.2f} €",
                f"{self.data.interest_rate:.2f}%",
                "0.00%",
                f"{self.data.interest_rate:.2f}%",
                self.data.years * 12,
                "",
                f"{self.results.monthly_payment_without_bonus:,.2f} €",
                f"{self.results.total_interest_without_bonus:,.2f} €",
                f"{self.results.total_paid_without_bonus:,.2f} €",
                "0.00 €",
                f"{self.results.total_paid_without_bonus:,.2f} €",
            ],
            "Con Bonificaciones": [
                f"{self.data.capital:,.2f} €",
                f"{self.data.interest_rate:.2f}%",
                f"{self.calculator.calculate_total_bonus():.2f}%",
                f"{max(0, self.data.interest_rate - self.calculator.calculate_total_bonus()):.2f}%",
                self.data.years * 12,
                "",
                f"{self.results.monthly_payment_with_bonus:,.2f} €",
                f"{self.results.total_interest_with_bonus:,.2f} €",
                f"{self.results.total_paid_with_bonus:,.2f} €",
                f"{self.results.total_bonus_costs:,.2f} €",
                f"{self.results.total_paid_with_bonus + self.results.total_bonus_costs:,.2f} €",
            ],
            "Diferencia": [
                "0.00 €",
                "0.00%",
                f"{self.calculator.calculate_total_bonus():.2f}%",
                f"{-self.calculator.calculate_total_bonus():.2f}%",
                "0",
                "",
                f"{self.results.monthly_payment_without_bonus - self.results.monthly_payment_with_bonus:,.2f} €",
                f"{self.results.total_interest_without_bonus - self.results.total_interest_with_bonus:,.2f} €",
                f"{self.results.total_paid_without_bonus - self.results.total_paid_with_bonus:,.2f} €",
                f"-{self.results.total_bonus_costs:,.2f} €",
                f"{self.results.real_savings:,.2f} €",
            ],
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Comparación", index=False)

    def _create_amortization_sheet(self, writer: pd.ExcelWriter, with_bonus: bool = False):
        """Crea la hoja con la tabla de amortización."""
        rate = self.data.interest_rate
        if with_bonus:
            rate = max(0, rate - self.calculator.calculate_total_bonus())

        schedule = self.calculator.calculate_amortization_schedule(rate)

        data = {
            "Mes": [s[0] for s in schedule],
            "Cuota (€)": [round(s[1], 2) for s in schedule],
            "Intereses (€)": [round(s[2], 2) for s in schedule],
            "Amortización (€)": [round(s[3], 2) for s in schedule],
            "Pendiente (€)": [round(s[4], 2) for s in schedule],
        }

        df = pd.DataFrame(data)
        sheet_name = "Amortización CON Bonif." if with_bonus else "Amortización SIN Bonif."
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _create_bonus_analysis_sheet(self, writer: pd.ExcelWriter):
        """Crea una hoja con análisis detallado de bonificaciones."""
        if not self.results:
            return

        months = self.data.years * 12
        yearly_card = self.data.card_annual_fee
        monthly_other = self.data.other_costs_monthly

        data = {
            "Bonificación": [
                "Domiciliación de nómina",
                "Seguro de vida",
                "Seguro de hogar",
                "Uso de tarjeta",
                "Otras bonificaciones",
                "",
                "TOTAL BONIFICACIONES",
            ],
            "Reducción de Tipo (%)": [
                f"{self.data.payroll_bonus}%",
                f"{self.data.life_insurance_bonus}%",
                f"{self.data.home_insurance_bonus}%",
                f"{self.data.card_bonus}%",
                f"{self.data.other_bonus}%",
                "",
                f"{self.calculator.calculate_total_bonus()}%",
            ],
            "Coste Mensual (€)": [
                0,
                self.data.life_insurance_cost_monthly,
                self.data.home_insurance_cost_monthly,
                yearly_card / 12,
                monthly_other,
                "",
                self.data.life_insurance_cost_monthly
                + self.data.home_insurance_cost_monthly
                + (yearly_card / 12)
                + monthly_other,
            ],
            "Coste Total (€)": [
                0,
                self.data.life_insurance_cost_monthly * months,
                self.data.home_insurance_cost_monthly * months,
                yearly_card * self.data.years,
                monthly_other * months,
                "",
                self.results.total_bonus_costs,
            ],
            "Ahorro en Intereses (€)": [
                "-",
                "-",
                "-",
                "-",
                "-",
                "",
                self.results.total_interest_without_bonus - self.results.total_interest_with_bonus,
            ],
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Análisis Bonificaciones", index=False)

    def _create_insurance_individual_analysis_sheet(self, writer: pd.ExcelWriter):
        """Crea una hoja con análisis individual de cada seguro."""
        if not self.results:
            return

        # Análisis seguro de vida
        life_analysis = self._analyze_individual_insurance(
            bonus_rate=self.data.life_insurance_bonus,
            monthly_cost=self.data.life_insurance_cost_monthly,
            insurance_name="Seguro de Vida",
        )

        # Análisis seguro de hogar
        home_analysis = self._analyze_individual_insurance(
            bonus_rate=self.data.home_insurance_bonus,
            monthly_cost=self.data.home_insurance_cost_monthly,
            insurance_name="Seguro de Hogar",
        )

        # Calcular punto de equilibrio para cada seguro
        life_breakeven = self._calculate_insurance_breakeven(self.data.life_insurance_bonus)
        home_breakeven = self._calculate_insurance_breakeven(self.data.home_insurance_bonus)

        data = {
            "Concepto": [
                "▼ SEGURO DE VIDA",
                "Bonificación aplicada (%)",
                "Coste mensual actual (€)",
                "Coste total durante hipoteca (€)",
                "Ahorro en intereses (€)",
                "Ahorro neto (€)",
                "¿Vale la pena?",
                "",
                "Análisis de rentabilidad:",
                "Coste mensual máximo rentable (€)",
                "Coste anual máximo rentable (€)",
                "Coste total máximo rentable (€)",
                "",
                "▼ SEGURO DE HOGAR",
                "Bonificación aplicada (%)",
                "Coste mensual actual (€)",
                "Coste total durante hipoteca (€)",
                "Ahorro en intereses (€)",
                "Ahorro neto (€)",
                "¿Vale la pena?",
                "",
                "Análisis de rentabilidad:",
                "Coste mensual máximo rentable (€)",
                "Coste anual máximo rentable (€)",
                "Coste total máximo rentable (€)",
                "",
                "▼ COMPARACIÓN",
                "Mejor seguro por rentabilidad",
                "Diferencia de ahorro (€)",
                "",
                "▼ RECOMENDACIONES",
                "Seguro de vida",
                "Seguro de hogar",
            ],
            "Valor": [
                "",
                f"{self.data.life_insurance_bonus}%",
                self.data.life_insurance_cost_monthly,
                life_analysis["total_cost"],
                life_analysis["interest_savings"],
                life_analysis["net_savings"],
                "SÍ ✓" if life_analysis["is_worth_it"] else "NO ✗",
                "",
                "",
                life_breakeven["max_monthly_cost"],
                life_breakeven["max_annual_cost"],
                life_breakeven["max_total_cost"],
                "",
                "",
                f"{self.data.home_insurance_bonus}%",
                self.data.home_insurance_cost_monthly,
                home_analysis["total_cost"],
                home_analysis["interest_savings"],
                home_analysis["net_savings"],
                "SÍ ✓" if home_analysis["is_worth_it"] else "NO ✗",
                "",
                "",
                home_breakeven["max_monthly_cost"],
                home_breakeven["max_annual_cost"],
                home_breakeven["max_total_cost"],
                "",
                "",
                self._get_best_insurance(life_analysis, home_analysis),
                abs(life_analysis["net_savings"] - home_analysis["net_savings"]),
                "",
                "",
                self._get_insurance_recommendation(
                    life_analysis, self.data.life_insurance_cost_monthly, life_breakeven
                ),
                self._get_insurance_recommendation(
                    home_analysis, self.data.home_insurance_cost_monthly, home_breakeven
                ),
            ],
        }

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Análisis Individual Seguros", index=False)

    def _analyze_individual_insurance(
        self, bonus_rate: float, monthly_cost: float, insurance_name: str
    ) -> dict:
        """Analiza la rentabilidad de un seguro individual."""
        # Crear datos temporales solo con este seguro
        temp_data = MortgageData(
            capital=self.data.capital,
            interest_rate=self.data.interest_rate,
            years=self.data.years,
            payroll_bonus=0.0,
            life_insurance_bonus=bonus_rate if "Vida" in insurance_name else 0.0,
            home_insurance_bonus=bonus_rate if "Hogar" in insurance_name else 0.0,
            card_bonus=0.0,
            other_bonus=0.0,
            life_insurance_cost_monthly=monthly_cost if "Vida" in insurance_name else 0.0,
            home_insurance_cost_monthly=monthly_cost if "Hogar" in insurance_name else 0.0,
            card_annual_fee=0.0,
            other_costs_monthly=0.0,
        )

        calculator = MortgageCalculator(temp_data)
        results = calculator.calculate()

        months = self.data.years * 12
        total_cost = monthly_cost * months
        interest_savings = results.total_interest_without_bonus - results.total_interest_with_bonus
        net_savings = interest_savings - total_cost

        return {
            "total_cost": total_cost,
            "interest_savings": interest_savings,
            "net_savings": net_savings,
            "is_worth_it": net_savings > 0,
            "monthly_cost": monthly_cost,
        }

    def _calculate_insurance_breakeven(self, bonus_rate: float) -> dict:
        """Calcula el coste máximo para que un seguro valga la pena."""
        if bonus_rate == 0:
            return {"max_monthly_cost": 0.0, "max_annual_cost": 0.0, "max_total_cost": 0.0}

        # Calcular ahorro en intereses con esta bonificación
        temp_data = MortgageData(
            capital=self.data.capital,
            interest_rate=self.data.interest_rate,
            years=self.data.years,
            payroll_bonus=0.0,
            life_insurance_bonus=bonus_rate,
            home_insurance_bonus=0.0,
            card_bonus=0.0,
            other_bonus=0.0,
            life_insurance_cost_monthly=0.0,
            home_insurance_cost_monthly=0.0,
            card_annual_fee=0.0,
            other_costs_monthly=0.0,
        )

        calculator = MortgageCalculator(temp_data)
        results = calculator.calculate()

        interest_savings = results.total_interest_without_bonus - results.total_interest_with_bonus
        months = self.data.years * 12

        max_total_cost = interest_savings
        max_monthly_cost = max_total_cost / months
        max_annual_cost = max_monthly_cost * 12

        return {
            "max_monthly_cost": round(max_monthly_cost, 2),
            "max_annual_cost": round(max_annual_cost, 2),
            "max_total_cost": round(max_total_cost, 2),
        }

    def _get_best_insurance(self, life_analysis: dict, home_analysis: dict) -> str:
        """Determina cuál seguro es más rentable."""
        if life_analysis["net_savings"] > home_analysis["net_savings"]:
            if life_analysis["is_worth_it"]:
                return "Seguro de Vida (más rentable)"
            else:
                return "Ninguno es rentable"
        elif home_analysis["net_savings"] > life_analysis["net_savings"]:
            if home_analysis["is_worth_it"]:
                return "Seguro de Hogar (más rentable)"
            else:
                return "Ninguno es rentable"
        else:
            if life_analysis["is_worth_it"]:
                return "Ambos igual de rentables"
            else:
                return "Ninguno es rentable"

    def _get_insurance_recommendation(
        self, analysis: dict, current_cost: float, breakeven: dict
    ) -> str:
        """Genera una recomendación para el seguro."""
        if analysis["is_worth_it"]:
            margin = breakeven["max_monthly_cost"] - current_cost
            margin_pct = (margin / breakeven["max_monthly_cost"]) * 100
            return f"✓ Contratar (margen: {margin:.2f}€/mes = {margin_pct:.1f}%)"
        else:
            excess = current_cost - breakeven["max_monthly_cost"]
            return f"✗ No contratar (excede punto equilibrio en {excess:.2f}€/mes)"

    def _apply_formatting(self, file_path: str):
        """Aplica formato visual al archivo Excel."""
        wb = load_workbook(file_path)

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        section_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        section_font = Font(bold=True, size=10)

        highlight_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        warning_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Formatear cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (AttributeError, TypeError):
                        pass
                adjusted_width = min(50, max(12, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width

            # Formatear encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Formatear filas con secciones (▼)
            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).startswith("▼"):
                    row[0].fill = section_fill
                    row[0].font = section_font

                # Resaltar la decisión principal
                if "vale la pena" in str(row[0].value).lower():
                    if "SÍ" in str(row[1].value):
                        row[1].fill = highlight_fill
                        row[1].font = Font(bold=True, size=12)
                    else:
                        row[1].fill = warning_fill
                        row[1].font = Font(bold=True, size=12)

        wb.save(file_path)

    def _add_formulas_to_sheets(self, file_path: str):
        """Añade fórmulas dinámicas a las hojas para que se actualicen automáticamente."""
        wb = load_workbook(file_path)

        # Formatear porcentajes en la hoja de entrada
        if "Datos de Entrada" in wb.sheetnames:
            ws = wb["Datos de Entrada"]
            # B3 = Interés, B7-B11 = Bonificaciones (dividir por 100 para formato %)
            percentage_cells = ["B3", "B7", "B8", "B9", "B10", "B11"]
            for cell_addr in percentage_cells:
                cell = ws[cell_addr]
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100  # Convertir a decimal
                    cell.number_format = "0.00%"

        # Referencias a celdas de entrada (hoja "Datos de Entrada")
        # B2 = Capital, B3 = Interés, B4 = Plazo
        # B7-B11 = Bonificaciones, B14-B17 = Costes

        # Hoja de Resumen
        if "Resumen" in wb.sheetnames:
            ws = wb["Resumen"]

            # Referencias a datos de entrada
            input_sheet = "'Datos de Entrada'"

            # Estructura de la hoja Resumen:
            # A1="Concepto", B1="Fórmula/Valor" (ENCABEZADO)
            # A2="▼ CÁLCULOS AUTOMÁTICOS", B2=vacío
            # A3="Meses totales", B3=fórmula -> debe leer de 'Datos de Entrada'!B5 (Plazo)
            # A4="Bonificación total (%)", B4=fórmula -> debe sumar B8:B12 de Datos de Entrada
            # A5="Tipo efectivo sin bonif. (%)", B5=fórmula -> debe leer B4 de Datos de Entrada
            # A6="Tipo efectivo con bonif. (%)", B6=fórmula -> B5-B4 (tipo - bonificaciones)
            # A7="Cuota mensual SIN bonificaciones (€)", B7=fórmula -> usa B5 (tipo sin bonif)
            # A8="Cuota mensual CON bonificaciones (€)", B8=fórmula -> usa B6 (tipo con bonif)
            # A9="Total a pagar SIN bonificaciones (€)", B9=fórmula
            # A10="Total a pagar CON bonificaciones (€)", B10=fórmula
            # A11="Intereses SIN bonificaciones (€)", B11=fórmula
            # A12="Intereses CON bonificaciones (€)", B12=fórmula
            # A13="Costes bonificaciones (€)", B13=fórmula
            # A14="Ahorro en intereses (€)", B14=fórmula
            # A15="Ahorro real (€)", B15=fórmula
            # A16="¿Vale la pena?", B16=fórmula
            # A17="Porcentaje de ahorro (%)", B17=fórmula

            # DATOS DE ENTRADA - Referencias:
            # B3 = Capital (fila 3)
            # B4 = Interés (fila 4)
            # B5 = Plazo años (fila 5)
            # B8 = Bonif nómina (fila 8)
            # B9 = Bonif vida (fila 9)
            # B10 = Bonif hogar (fila 10)
            # B11 = Bonif tarjeta (fila 11)
            # B12 = Otras bonif (fila 12)
            # B15 = Coste vida (fila 15)
            # B16 = Coste hogar (fila 16)
            # B17 = Coste tarjeta anual (fila 17)
            # B18 = Otros costes (fila 18)

            # Calcular meses totales (fila 3) -> Plazo * 12
            ws["B3"] = f"={input_sheet}!B5*12"
            ws["B3"].number_format = "0"

            # Total bonificaciones (fila 4) -> Suma B8+B9+B10+B11+B12
            ws["B4"] = (
                f"={input_sheet}!B8+{input_sheet}!B9+{input_sheet}!B10+{input_sheet}!B11+{input_sheet}!B12"
            )
            ws["B4"].number_format = "0.00%"

            # Tipo efectivo sin bonificaciones (fila 5) -> Interés de B4
            ws["B5"] = f"={input_sheet}!B4"
            ws["B5"].number_format = "0.00%"

            # Tipo efectivo con bonificaciones (fila 6) -> B5 - B4 (tipo - bonificaciones)
            ws["B6"] = "=MAX(0, B5-B4)"
            ws["B6"].number_format = "0.00%"

            # Cuota mensual SIN bonificaciones (fila 7) -> Capital en B3, usa tipo B5, meses B3
            ws["B7"] = (
                f"=IF(B5=0, {input_sheet}!B3/B3, {input_sheet}!B3*(B5/12)*(1+B5/12)^B3/((1+B5/12)^B3-1))"
            )
            ws["B7"].number_format = "#,##0.00"

            # Cuota mensual CON bonificaciones (fila 8) -> Capital en B3, usa tipo B6, meses B3
            ws["B8"] = (
                f"=IF(B6=0, {input_sheet}!B3/B3, {input_sheet}!B3*(B6/12)*(1+B6/12)^B3/((1+B6/12)^B3-1))"
            )
            ws["B8"].number_format = "#,##0.00"

            # Total a pagar sin bonificaciones (fila 9) -> B7 * B3
            ws["B9"] = "=B7*B3"
            ws["B9"].number_format = "#,##0.00"

            # Total a pagar con bonificaciones (fila 10) -> B8 * B3
            ws["B10"] = "=B8*B3"
            ws["B10"].number_format = "#,##0.00"

            # Intereses sin bonificaciones (fila 11) -> B9 - Capital (B3 de Datos)
            ws["B11"] = f"=B9-{input_sheet}!B3"
            ws["B11"].number_format = "#,##0.00"

            # Intereses con bonificaciones (fila 12) -> B10 - Capital (B3 de Datos)
            ws["B12"] = f"=B10-{input_sheet}!B3"
            ws["B12"].number_format = "#,##0.00"

            # Costes de bonificaciones (fila 13) -> (B15+B16+B18)*meses + B17*años
            ws["B13"] = (
                f"=({input_sheet}!B15+{input_sheet}!B16+{input_sheet}!B18)*B3+{input_sheet}!B17*{input_sheet}!B5"
            )
            ws["B13"].number_format = "#,##0.00"

            # Ahorro nominal (fila 14) -> B11 - B12
            ws["B14"] = "=B11-B12"
            ws["B14"].number_format = "#,##0.00"

            # Ahorro real (fila 15) -> B14 - B13
            ws["B15"] = "=B14-B13"
            ws["B15"].number_format = "#,##0.00"

            # ¿Vale la pena? (fila 16) -> Si B15 > 0
            ws["B16"] = '=IF(B15>0,"SÍ ✓","NO ✗")'

            # Porcentaje de ahorro (fila 17) -> (B15 / B9) * 100
            ws["B17"] = "=IF(B9=0,0,(B15/B9)*100)"
            ws["B17"].number_format = "0.00"

        # Hoja de Análisis de Bonificaciones
        if "Análisis Bonificaciones" in wb.sheetnames:
            ws = wb["Análisis Bonificaciones"]
            input_sheet = "'Datos de Entrada'"

            # Total bonificaciones -> B8+B9+B10+B11+B12
            ws["B7"] = (
                f"={input_sheet}!B8+{input_sheet}!B9+{input_sheet}!B10+{input_sheet}!B11+{input_sheet}!B12"
            )

            # Coste mensual total -> B15 (vida) + B16 (hogar) + B17/12 (tarjeta) + B18 (otros)
            ws["C7"] = (
                f"={input_sheet}!B15+{input_sheet}!B16+{input_sheet}!B17/12+{input_sheet}!B18"
            )
            ws["C7"].number_format = "#,##0.00"

            # Coste total -> (B15+B16+B18)*Plazo*12 + B17*Plazo
            ws["D7"] = (
                f"=({input_sheet}!B15+{input_sheet}!B16+{input_sheet}!B18)*{input_sheet}!B5*12+{input_sheet}!B17*{input_sheet}!B5"
            )
            ws["D7"].number_format = "#,##0.00"

            # Ahorro en intereses (referencia a la hoja Resumen, B14 = ahorro nominal)
            ws["E7"] = "=Resumen!B14"
            ws["E7"].number_format = "#,##0.00"

        # Hoja de Análisis Individual Seguros
        if "Análisis Individual Seguros" in wb.sheetnames:
            ws = wb["Análisis Individual Seguros"]
            input_sheet = "'Datos de Entrada'"

            # Seguro de Vida
            # Bonificación -> B9 (fila 9)
            ws["B2"] = f"={input_sheet}!B9"
            # Coste mensual -> B15 (fila 15)
            ws["B3"] = f"={input_sheet}!B15"
            # Coste total -> B15 * B5 (Plazo) * 12
            ws["B4"] = f"={input_sheet}!B15*{input_sheet}!B5*12"
            ws["B4"].number_format = "#,##0.00"

            # Seguro de Hogar
            # Bonificación -> B10 (fila 10)
            ws["B15"] = f"={input_sheet}!B10"
            # Coste mensual -> B16 (fila 16)
            ws["B16"] = f"={input_sheet}!B16"
            # Coste total -> B16 * B5 (Plazo) * 12
            ws["B17"] = f"={input_sheet}!B16*{input_sheet}!B5*12"
            ws["B17"].number_format = "#,##0.00"

        wb.save(file_path)
